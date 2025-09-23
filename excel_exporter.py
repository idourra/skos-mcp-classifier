#!/usr/bin/env python3
# excel_exporter.py - Exportador completo para Excel
import json
from datetime import datetime
from client.classify_standard_api import classify
from utils.export_config import get_full_export_path, ensure_export_structure

def export_to_excel(products_data, filename=None):
    """
    Exporta productos clasificados a Excel con formato profesional
    
    Args:
        products_data: Lista de dicts con 'text'/'product' e 'id'/'sku' (opcional)
        filename: Nombre del archivo (opcional, se genera automático si no se proporciona)
    
    Returns:
        str: Nombre del archivo generado
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("❌ Error: Se necesita openpyxl para Excel")
        print("💡 Instalar con: pip install openpyxl")
        return None
    
    # Asegurar que existe la estructura de directorios
    ensure_export_structure()
    
    # Generar ruta completa si no se proporciona
    if not filename:
        output_path = get_full_export_path("productos_clasificados", "excel")
        filename = str(output_path)
    
    # Crear workbook y worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos Clasificados"
    
    # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "ID Producto",
        "Descripción del Producto", 
        "Categoría SKOS",
        "Notación SKOS",
        "URI Concepto",
        "Nivel de Confianza",
        "Timestamp Clasificación"
    ]
    
    # Escribir headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    print(f"📊 Clasificando {len(products_data)} productos...")
    
    # Procesar cada producto
    results = []
    for row_idx, product in enumerate(products_data, 2):
        # Extraer texto e ID
        if isinstance(product, dict):
            product_text = product.get('text', product.get('product', ''))
            product_id = product.get('id', product.get('sku', ''))
        else:
            product_text = str(product)
            product_id = f"PROD-{row_idx-1:03d}"
        
        print(f"  [{row_idx-1}/{len(products_data)}] Clasificando: {product_text}")
        
        try:
            # Clasificar producto
            result = classify(product_text, product_id)
            
            # Escribir datos en Excel
            ws.cell(row=row_idx, column=1, value=product_id)
            ws.cell(row=row_idx, column=2, value=product_text)
            ws.cell(row=row_idx, column=3, value=result.get('prefLabel', 'N/A'))
            ws.cell(row=row_idx, column=4, value=result.get('notation', 'N/A'))
            ws.cell(row=row_idx, column=5, value=result.get('concept_uri', 'N/A'))
            ws.cell(row=row_idx, column=6, value=result.get('confidence', 0))
            ws.cell(row=row_idx, column=7, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            
            # Aplicar bordes
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).border = border
            
            # Guardar resultado para stats
            results.append({
                'product_id': product_id,
                'product_text': product_text,
                'classification': result,
                'success': True
            })
            
        except Exception as e:
            print(f"    ❌ Error clasificando '{product_text}': {e}")
            # Escribir error en Excel
            ws.cell(row=row_idx, column=1, value=product_id)
            ws.cell(row=row_idx, column=2, value=product_text)
            ws.cell(row=row_idx, column=3, value=f"ERROR: {str(e)}")
            
            results.append({
                'product_id': product_id,
                'product_text': product_text,
                'error': str(e),
                'success': False
            })
    
    # Ajustar ancho de columnas
    column_widths = [15, 40, 25, 15, 50, 15, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Crear hoja de resumen
    summary_ws = wb.create_sheet("Resumen")
    
    # Stats del resumen
    successful = sum(1 for r in results if r['success'])
    failed = len(results) - successful
    
    summary_data = [
        ["📊 RESUMEN DE CLASIFICACIÓN", ""],
        ["", ""],
        ["Total de productos:", len(results)],
        ["Clasificaciones exitosas:", successful],
        ["Errores:", failed],
        ["Tasa de éxito:", f"{(successful/len(results)*100):.1f}%" if results else "0%"],
        ["", ""],
        ["Fecha de procesamiento:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]
    
    for row_idx, (label, value) in enumerate(summary_data, 1):
        summary_ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        summary_ws.cell(row=row_idx, column=2, value=value)
    
    # Categorías encontradas
    categories = {}
    for result in results:
        if result['success']:
            category = result['classification'].get('prefLabel', 'N/A')
            categories[category] = categories.get(category, 0) + 1
    
    if categories:
        summary_ws.cell(row=len(summary_data) + 2, column=1, value="📂 Categorías encontradas:").font = Font(bold=True)
        for idx, (category, count) in enumerate(sorted(categories.items()), len(summary_data) + 3):
            summary_ws.cell(row=idx, column=1, value=f"  • {category}")
            summary_ws.cell(row=idx, column=2, value=f"{count} productos")
    
    # Ajustar anchos del resumen
    summary_ws.column_dimensions['A'].width = 30
    summary_ws.column_dimensions['B'].width = 20
    
    # Guardar archivo
    wb.save(filename)
    
    print(f"\n✅ Excel generado exitosamente:")
    print(f"   📁 Archivo: {filename}")
    print(f"   📊 Productos procesados: {len(results)}")
    print(f"   ✅ Exitosos: {successful}")
    if failed > 0:
        print(f"   ❌ Errores: {failed}")
    
    return filename, results

def main():
    """Ejemplo de uso del exportador a Excel"""
    
    # Productos de ejemplo
    sample_products = [
        {"text": "yogur natural griego 0% grasa 500g", "id": "YOG-001"},
        {"text": "pan integral de centeno rebanado", "id": "PAN-002"},
        {"text": "aceite de oliva extra virgen 1L", "id": "ACE-003"},
        {"text": "queso manchego curado 250g", "id": "QUE-004"},
        {"text": "miel de abeja multifloral 500ml", "id": "MIE-005"},
        {"text": "arroz basmati integral 1kg", "id": "ARR-006"},
        {"text": "cerveza artesanal IPA 355ml", "id": "BEB-007"},
        {"text": "salmón fresco filete 200g", "id": "PES-008"},
        {"text": "chocolate negro 70% cacao", "id": "CHO-009"},
        {"text": "agua mineral con gas 500ml", "id": "AGU-010"}
    ]
    
    print("🚀 Exportador a Excel - Ejemplo completo")
    print("=" * 50)
    
    filename, results = export_to_excel(sample_products)
    
    if filename:
        print(f"\n🎉 ¡Proceso completado!")
        print(f"📄 Archivo Excel: {filename}")
        print("📋 El archivo contiene:")
        print("   • Hoja 'Productos Clasificados' con todos los datos")
        print("   • Hoja 'Resumen' con estadísticas")
        print("   • Formato profesional con colores y bordes")

if __name__ == "__main__":
    main()