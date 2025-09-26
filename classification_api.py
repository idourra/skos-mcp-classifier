#!/usr/bin/env python3
# classification_api.py - API REST completa para clasificación SKOS con múltiples taxonomías
from fastapi import FastAPI, HTTPException, BackgroundTasks, Query
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from enum import Enum
import uuid
import os
import csv
from pathlib import Path
from datetime import datetime
from client.classify_standard_api import classify
from core.non_classifiable_handler import enhance_classification_error_handling
from utils.export_config import get_full_export_path, ensure_export_structure, EXPORTS_BASE_DIR
from server.taxonomy_endpoints import taxonomy_router

app = FastAPI(
    title="Multi-Taxonomy SKOS Product Classifier API",
    description="API REST para clasificar productos usando múltiples taxonomías SKOS",
    version="2.0.0"
)

# Incluir router de taxonomías
app.include_router(taxonomy_router)

# Modelos Pydantic
class ProductRequest(BaseModel):
    text: str = Field(..., description="Descripción del producto a clasificar")
    product_id: Optional[str] = Field(None, description="ID/SKU opcional del producto")

class BatchProductRequest(BaseModel):
    products: List[ProductRequest] = Field(..., description="Lista de productos a clasificar")

# Modelos para información de costos OpenAI
class OpenAIUsage(BaseModel):
    prompt_tokens: int = Field(..., description="Tokens utilizados en el prompt")
    completion_tokens: int = Field(..., description="Tokens generados en la respuesta") 
    total_tokens: int = Field(..., description="Total de tokens utilizados")

class OpenAICostUSD(BaseModel):
    prompt: float = Field(..., description="Costo del prompt en USD")
    completion: float = Field(..., description="Costo de la respuesta en USD")
    total: float = Field(..., description="Costo total en USD")

class OpenAICostBreakdown(BaseModel):
    base_model_for_pricing: str = Field(..., description="Modelo base usado para el pricing")
    prompt_cost_per_1m_tokens: float = Field(..., description="Costo por 1M tokens de prompt")
    completion_cost_per_1m_tokens: float = Field(..., description="Costo por 1M tokens de completion")
    calculation_timestamp: str = Field(..., description="Timestamp del cálculo")

class OpenAICostInfo(BaseModel):
    model: str = Field(..., description="Modelo exacto utilizado por OpenAI")
    usage: OpenAIUsage = Field(..., description="Información de tokens utilizados")
    cost_usd: OpenAICostUSD = Field(..., description="Costos en USD")
    cost_breakdown: OpenAICostBreakdown = Field(..., description="Desglose detallado de costos")
    api_calls: int = Field(..., description="Número de llamadas a la API de OpenAI")

# Modelo para información de taxonomía
class TaxonomyInfo(BaseModel):
    id: str = Field(..., description="ID de la taxonomía utilizada")
    name: str = Field(..., description="Nombre de la taxonomía")
    is_default: bool = Field(..., description="Si se usó la taxonomía por defecto")

# Modelos para el endpoint unificado
class UnifiedProductRequest(BaseModel):
    products: List[ProductRequest] = Field(..., description="Lista de productos a clasificar (1 o más)", min_items=1)
    
class UnifiedClassificationResponse(BaseModel):
    total: int = Field(..., description="Total de productos procesados")
    successful: int = Field(..., description="Productos clasificados exitosamente") 
    failed: int = Field(..., description="Productos que fallaron")
    results: List[Dict[str, Any]] = Field(..., description="Array con resultados de clasificación")
    processing_time_seconds: Optional[float] = Field(None, description="Tiempo de procesamiento en segundos")
    timestamp: str = Field(..., description="Timestamp del procesamiento")
    openai_cost_info: Optional[OpenAICostInfo] = Field(None, description="Información de costos de OpenAI")
    taxonomy_used: Optional[TaxonomyInfo] = Field(None, description="Información de la taxonomía utilizada")

class ExportRequest(BaseModel):
    products: List[ProductRequest] = Field(..., description="Lista de productos para exportar")
    format: str = Field(..., description="Formato de exportación: 'csv' o 'excel'")
    filename: Optional[str] = Field(None, description="Nombre del archivo (opcional)")
    
class ClassificationResponse(BaseModel):
    product_id: Optional[str]
    search_text: str
    concept_uri: str
    prefLabel: str
    notation: str
    level: Optional[int]
    confidence: float
    timestamp: str
    taxonomy_used: Optional[TaxonomyInfo] = Field(None, description="Información de la taxonomía utilizada")
    
class BatchClassificationResponse(BaseModel):
    total: int
    successful: int 
    failed: int
    results: List[Dict[str, Any]]
    batch_id: str

class ExportResponse(BaseModel):
    status: str
    filename: str
    download_url: str
    total_products: int
    successful: int
    failed: int
    file_size_bytes: Optional[int] = None
    timestamp: str
    
class ErrorResponse(BaseModel):
    error: str
    detail: Optional[str] = None
    timestamp: str

# Modelos para clasificación asíncrona
class JobStatus(str, Enum):
    """Estados posibles de un job asíncrono"""
    QUEUED = "queued"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"
    CANCELLED = "cancelled"

class AsyncClassificationRequest(BaseModel):
    """Request para clasificación asíncrona"""
    products: List[ProductRequest] = Field(..., description="Lista de productos a clasificar", min_items=1)
    callback_url: Optional[str] = Field(None, description="URL opcional para notificación cuando complete")
    priority: Optional[int] = Field(1, description="Prioridad del job (1=alta, 5=baja)", ge=1, le=5)
    
class JobProgress(BaseModel):
    """Progreso de un job"""
    current: int = Field(..., description="Productos procesados")
    total: int = Field(..., description="Total de productos")
    percentage: float = Field(..., description="Porcentaje completado")

class AsyncJobResponse(BaseModel):
    """Respuesta al crear un job asíncrono"""
    job_id: str = Field(..., description="ID único del job")
    status: JobStatus = Field(..., description="Estado actual del job")
    message: str = Field(..., description="Mensaje descriptivo")
    total_products: int = Field(..., description="Total de productos a procesar")
    estimated_completion_time: Optional[str] = Field(None, description="Estimación de finalización")
    created_at: str = Field(..., description="Timestamp de creación")
    status_url: str = Field(..., description="URL para consultar estado")
    result_url: str = Field(..., description="URL para obtener resultados")

class JobStatusResponse(BaseModel):
    """Respuesta del estado de un job"""
    job_id: str = Field(..., description="ID del job")
    status: JobStatus = Field(..., description="Estado actual")
    progress: Optional[JobProgress] = Field(None, description="Progreso del procesamiento")
    created_at: str = Field(..., description="Timestamp de creación")
    started_at: Optional[str] = Field(None, description="Timestamp de inicio")
    completed_at: Optional[str] = Field(None, description="Timestamp de finalización")
    error_message: Optional[str] = Field(None, description="Mensaje de error si falló")
    total_products: int = Field(..., description="Total de productos")
    estimated_completion_time: Optional[str] = Field(None, description="Estimación de finalización")

class JobResultResponse(BaseModel):
    """Respuesta con resultados de un job completado"""
    job_id: str = Field(..., description="ID del job")
    status: JobStatus = Field(..., description="Estado del job")
    total: int = Field(..., description="Total de productos procesados")
    successful: int = Field(..., description="Productos clasificados exitosamente")
    failed: int = Field(..., description="Productos que fallaron")
    results: List[Dict[str, Any]] = Field(..., description="Resultados de clasificación")
    processing_time_seconds: Optional[float] = Field(None, description="Tiempo total de procesamiento")
    created_at: str = Field(..., description="Timestamp de creación")
    started_at: Optional[str] = Field(None, description="Timestamp de inicio")
    completed_at: Optional[str] = Field(None, description="Timestamp de finalización")
    openai_cost_info: Optional[OpenAICostInfo] = Field(None, description="Información de costos de OpenAI")

# Store para trabajos en background (en producción usar Redis/DB)
background_jobs = {}

# Helper functions para jobs asíncronos
def estimate_completion_time(num_products: int, avg_time_per_product: float = 1.5) -> str:
    """Estimar tiempo de finalización basado en número de productos"""
    total_seconds = num_products * avg_time_per_product
    completion_time = datetime.now().timestamp() + total_seconds
    return datetime.fromtimestamp(completion_time).isoformat()

def create_job_metadata(products: List[ProductRequest], priority: int = 1) -> dict:
    """Crear metadata inicial para un job"""
    now = datetime.now().isoformat()
    return {
        "status": JobStatus.QUEUED,
        "created_at": now,
        "started_at": None,
        "completed_at": None,
        "total_products": len(products),
        "priority": priority,
        "progress": {"current": 0, "total": len(products), "percentage": 0.0},
        "estimated_completion_time": estimate_completion_time(len(products)),
        "error_message": None,
        "results": [],
        "processing_time_seconds": None,
        "openai_cost_info": None
    }

@app.get("/")
def root():
    """Endpoint raíz con información de la API"""
    return {
        "message": "SKOS Product Classifier API",
        "version": "2.0.0",
        "description": "API REST para clasificación de productos usando taxonomía SKOS",
        "endpoints": {
            "classification": {
                "classify_products": {
                    "url": "/classify/products",
                    "method": "POST",
                    "description": "Clasifica 1 o N productos en una sola llamada",
                    "example_single": {
                        "products": [
                            {"text": "leche descremada", "product_id": "SKU001"}
                        ]
                    },
                    "example_multiple": {
                        "products": [
                            {"text": "arroz blanco", "product_id": "SKU001"},
                            {"text": "pollo congelado", "product_id": "SKU002"},
                            {"text": "yogurt natural", "product_id": "SKU003"}
                        ]
                    }
                }
            },
            "export": {
                "export_csv": {
                    "url": "/export/csv",
                    "method": "POST", 
                    "description": "Exportar productos clasificados a CSV"
                },
                "export_excel": {
                    "url": "/export/excel",
                    "method": "POST",
                    "description": "Exportar productos clasificados a Excel"
                },
                "download": {
                    "url": "/download/{filename}",
                    "method": "GET",
                    "description": "Descargar archivo exportado"
                }
            },
            "system": {
                "health": {
                    "url": "/health",
                    "method": "GET",
                    "description": "Verificar estado del sistema"
                },
                "stats": {
                    "url": "/stats", 
                    "method": "GET",
                    "description": "Estadísticas de uso de la API"
                }
            }
        }
    }

@app.get("/health")
def health_check():
    """Health check endpoint"""
    try:
        # Prueba rápida de clasificación
        classify("test product")
        return {
            "status": "healthy",
            "mcp_server": "connected",
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        return {
            "status": "unhealthy", 
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }

@app.post("/classify", response_model=ClassificationResponse)
def classify_single_product(
    request: ProductRequest,
    taxonomy: Optional[str] = Query(None, description="ID de taxonomía específica a usar (opcional, usa default si no se especifica)")
):
    """
    Clasificar un producto individual con soporte para múltiples taxonomías.
    
    **Funcionalidad mejorada:**
    - **taxonomy**: Parámetro opcional para especificar taxonomía específica
    - Si no se especifica taxonomy, usa la taxonomía configurada como default
    - Respuesta incluye información detallada de costos OpenAI
    
    **Taxonomías disponibles:** Consulta `/taxonomies` para ver opciones disponibles.
    
    **Ejemplos de uso:**
    - Sin taxonomía específica: usa la por defecto (treew-skos)
    - Con taxonomía: `?taxonomy=electronics-taxonomy`
    """
    try:
        # Validar y usar taxonomía (por defecto si no se especifica)
        from utils.taxonomy_config import validate_taxonomy_id, get_taxonomy_info
        
        validated_taxonomy = validate_taxonomy_id(taxonomy)
        taxonomy_info = get_taxonomy_info(validated_taxonomy)
        
        result = classify(request.text, request.product_id, validated_taxonomy)
        
        # Verificar que el resultado tenga los campos necesarios
        if 'error' in result:
            raise HTTPException(
                status_code=422,
                detail=f"Error en clasificación: {result.get('error', 'Error desconocido')}"
            )
        
        # Añadir información de la taxonomía utilizada
        result['taxonomy_used'] = {
            'id': validated_taxonomy,
            'name': taxonomy_info.get('name', 'Unknown'),
            'is_default': taxonomy is None  # True si usó la por defecto
        }
        
        return ClassificationResponse(
            product_id=result.get('product_id'),
            search_text=result.get('search_text', request.text),
            concept_uri=result.get('concept_uri', ''),
            prefLabel=result.get('prefLabel', ''),
            notation=result.get('notation', ''),
            level=result.get('level'),
            confidence=result.get('confidence', 0.0),
            timestamp=datetime.now().isoformat(),
            taxonomy_used=TaxonomyInfo(
                id=validated_taxonomy,
                name=taxonomy_info.get('name', 'Unknown'),
                is_default=taxonomy is None
            )
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/classify/products", response_model=UnifiedClassificationResponse)
def classify_products_unified(
    request: UnifiedProductRequest,
    taxonomy: Optional[str] = Query(None, description="ID de taxonomía específica a usar (opcional)")
):
    """
    Endpoint unificado para clasificar productos con soporte multi-taxonomía.
    
    Acepta un array de productos (1 o más) y devuelve un array con los resultados.
    - Para clasificar 1 producto: envía array con 1 elemento
    - Para clasificar N productos: envía array con N elementos
    
    **Nuevo**: Soporte para múltiples taxonomías
    - **taxonomy**: ID de taxonomía específica (opcional)
    - Si no se especifica, usa la taxonomía por defecto
    - Las taxonomías disponibles se pueden consultar en /taxonomies
    
    Respuesta siempre en formato array con todos los resultados de clasificación.
    Incluye información agregada de costos de OpenAI para toda la operación.
    """
    import time
    start_time = time.time()
    
    # Validar que el array no esté vacío
    if not request.products:
        raise HTTPException(
            status_code=422,
            detail="Array de productos no puede estar vacío. Debe contener al menos 1 producto."
        )
    
    results = []
    successful = 0
    failed = 0
    
    # Variables para consolidar información de costos
    total_prompt_tokens = 0
    total_completion_tokens = 0
    total_cost_usd = 0.0
    total_api_calls = 0
    model_used = None
    cost_info_sample = None
    
    # Validar y configurar taxonomía (una vez para toda la operación)
    from utils.taxonomy_config import validate_taxonomy_id, get_taxonomy_info
    
    validated_taxonomy = validate_taxonomy_id(taxonomy)
    taxonomy_info = get_taxonomy_info(validated_taxonomy)
    
    for idx, product in enumerate(request.products):
        try:
            result = classify(product.text, product.product_id, validated_taxonomy)
            
            # Extraer información de costos si está disponible
            if 'openai_cost' in result:
                cost_data = result['openai_cost']
                if 'usage' in cost_data:
                    total_prompt_tokens += cost_data['usage'].get('prompt_tokens', 0)
                    total_completion_tokens += cost_data['usage'].get('completion_tokens', 0)
                if 'cost_usd' in cost_data:
                    total_cost_usd += cost_data['cost_usd'].get('total', 0.0)
                if 'api_calls' in cost_data:
                    total_api_calls += cost_data['api_calls']
                if 'model' in cost_data:
                    model_used = cost_data['model']  # Usar el último modelo
                if not cost_info_sample:
                    cost_info_sample = cost_data  # Guardar una muestra para los breakdowns
            
            if 'error' not in result:
                # Formato unificado de respuesta exitosa
                classification_result = {
                    "index": idx,
                    "product_id": product.product_id,
                    "search_text": product.text,
                    "concept_uri": result.get('concept_uri', ''),
                    "prefLabel": result.get('prefLabel', ''),
                    "notation": result.get('notation', ''),
                    "level": result.get('level'),
                    "confidence": result.get('confidence', 0.0),
                    "status": "success",
                    "timestamp": datetime.now().isoformat()
                }
                results.append(classification_result)
                successful += 1
            else:
                # Formato unificado de respuesta con error
                error_result = {
                    "index": idx,
                    "product_id": product.product_id,
                    "search_text": product.text,
                    "error": result['error'],
                    "status": "error",
                    "timestamp": datetime.now().isoformat()
                }
                results.append(error_result)
                failed += 1
                
        except Exception as e:
            # Formato unificado para excepciones
            exception_result = {
                "index": idx,
                "product_id": product.product_id,
                "search_text": product.text,
                "error": str(e),
                "status": "exception",
                "timestamp": datetime.now().isoformat()
            }
            results.append(exception_result)
            failed += 1
    
    processing_time = time.time() - start_time
    
    # Crear información consolidada de costos OpenAI
    openai_cost_info = None
    if cost_info_sample and model_used:
        openai_cost_info = OpenAICostInfo(
            model=model_used,
            usage=OpenAIUsage(
                prompt_tokens=total_prompt_tokens,
                completion_tokens=total_completion_tokens,
                total_tokens=total_prompt_tokens + total_completion_tokens
            ),
            cost_usd=OpenAICostUSD(
                prompt=round(total_cost_usd * (total_prompt_tokens / (total_prompt_tokens + total_completion_tokens)) if total_prompt_tokens + total_completion_tokens > 0 else 0, 6),
                completion=round(total_cost_usd * (total_completion_tokens / (total_prompt_tokens + total_completion_tokens)) if total_prompt_tokens + total_completion_tokens > 0 else 0, 6),
                total=round(total_cost_usd, 6)
            ),
            cost_breakdown=OpenAICostBreakdown(
                base_model_for_pricing=cost_info_sample.get('cost_breakdown', {}).get('base_model_for_pricing', model_used),
                prompt_cost_per_1m_tokens=cost_info_sample.get('cost_breakdown', {}).get('prompt_cost_per_1m_tokens', 0.0),
                completion_cost_per_1m_tokens=cost_info_sample.get('cost_breakdown', {}).get('completion_cost_per_1m_tokens', 0.0),
                calculation_timestamp=datetime.now().isoformat()
            ),
            api_calls=total_api_calls
        )
    
    return UnifiedClassificationResponse(
        total=len(request.products),
        successful=successful,
        failed=failed,
        results=results,
        processing_time_seconds=round(processing_time, 3),
        timestamp=datetime.now().isoformat(),
        openai_cost_info=openai_cost_info,
        taxonomy_used=TaxonomyInfo(
            id=validated_taxonomy,
            name=taxonomy_info.get('name', 'Unknown'),
            is_default=taxonomy is None
        )
    )

# === ENDPOINTS ASÍNCRONOS ===

@app.post("/classify/async", response_model=AsyncJobResponse, 
          summary="Iniciar clasificación asíncrona",
          description="Crear un job de clasificación que se procesa en background")
def create_async_classification_job(request: AsyncClassificationRequest, background_tasks: BackgroundTasks):
    """
    **Clasificación Asíncrona de Productos**
    
    Crea un job de clasificación que se ejecuta en background, ideal para lotes grandes.
    
    **Características:**
    - Procesamiento no bloqueante
    - Tracking de progreso en tiempo real
    - Estimación de tiempo de finalización
    - Soporte para prioridades
    - Callback opcional cuando complete
    
    **Flujo de trabajo:**
    1. POST /classify/async → Recibe job_id
    2. GET /classify/status/{job_id} → Consultar progreso
    3. GET /classify/result/{job_id} → Obtener resultados finales
    
    **Prioridades:**
    - 1: Alta prioridad (procesamiento más rápido)
    - 5: Baja prioridad (puede tomar más tiempo)
    """
    # Generar ID único para el job
    job_id = str(uuid.uuid4())
    
    # Crear metadata inicial del job
    job_metadata = create_job_metadata(request.products, request.priority or 1)
    background_jobs[job_id] = job_metadata
    
    # Agregar tarea de procesamiento en background
    background_tasks.add_task(process_async_classification, request.products, job_id, request.callback_url)
    
    return AsyncJobResponse(
        job_id=job_id,
        status=JobStatus.QUEUED,
        message=f"Job de clasificación creado exitosamente. Procesando {len(request.products)} productos.",
        total_products=len(request.products),
        estimated_completion_time=job_metadata["estimated_completion_time"],
        created_at=job_metadata["created_at"],
        status_url=f"/classify/status/{job_id}",
        result_url=f"/classify/result/{job_id}"
    )

@app.get("/classify/status/{job_id}", response_model=JobStatusResponse,
         summary="Consultar estado de job",
         description="Obtener estado actual y progreso de un job de clasificación")
def get_classification_job_status(job_id: str):
    """
    **Consultar Estado de Job Asíncrono**
    
    Obtiene el estado actual, progreso y metadata de un job de clasificación.
    
    **Estados posibles:**
    - `queued`: En cola, esperando procesamiento
    - `processing`: Ejecutándose actualmente
    - `completed`: Finalizado exitosamente
    - `failed`: Falló por error
    - `cancelled`: Cancelado por el usuario
    
    **Información incluida:**
    - Progreso actual (productos procesados)
    - Porcentaje de completitud
    - Tiempo estimado de finalización
    - Timestamps de creación, inicio y finalización
    """
    if job_id not in background_jobs:
        raise HTTPException(
            status_code=404, 
            detail=f"Job {job_id} no encontrado. Verifique el job_id o que no haya expirado."
        )
    
    job_data = background_jobs[job_id]
    
    return JobStatusResponse(
        job_id=job_id,
        status=JobStatus(job_data["status"]),
        progress=JobProgress(
            current=job_data["progress"]["current"],
            total=job_data["progress"]["total"],
            percentage=job_data["progress"]["percentage"]
        ) if job_data["status"] in [JobStatus.PROCESSING, JobStatus.COMPLETED] else None,
        created_at=job_data["created_at"],
        started_at=job_data.get("started_at"),
        completed_at=job_data.get("completed_at"),
        error_message=job_data.get("error_message"),
        total_products=job_data["total_products"],
        estimated_completion_time=job_data.get("estimated_completion_time")
    )

@app.get("/classify/result/{job_id}", response_model=JobResultResponse,
         summary="Obtener resultados de job",
         description="Obtener resultados finales de un job completado")
def get_classification_job_result(job_id: str):
    """
    **Obtener Resultados de Job Completado**
    
    Retorna los resultados finales de un job de clasificación que ha terminado.
    
    **Requisitos:**
    - El job debe estar en estado `completed`
    - Solo jobs finalizados exitosamente tienen resultados
    
    **Contenido de respuesta:**
    - Todos los productos clasificados
    - Resumen de éxitos y fallos
    - Información de costos de OpenAI
    - Tiempos de procesamiento
    - Timestamps completos
    
    **Nota:** Los resultados se mantienen en memoria por tiempo limitado.
    En producción se recomienda usar base de datos persistente.
    """
    if job_id not in background_jobs:
        raise HTTPException(
            status_code=404, 
            detail=f"Job {job_id} no encontrado. Verifique el job_id o que no haya expirado."
        )
    
    job_data = background_jobs[job_id]
    
    if job_data["status"] != JobStatus.COMPLETED:
        raise HTTPException(
            status_code=409,
            detail=f"Job {job_id} no está completado. Estado actual: {job_data['status']}. Use /classify/status/{job_id} para consultar progreso."
        )
    
    return JobResultResponse(
        job_id=job_id,
        status=JobStatus(job_data["status"]),
        total=job_data["total_products"],
        successful=len([r for r in job_data["results"] if r.get("status") == "success"]),
        failed=len([r for r in job_data["results"] if r.get("status") == "failed"]),
        results=job_data["results"],
        processing_time_seconds=job_data.get("processing_time_seconds"),
        created_at=job_data["created_at"],
        started_at=job_data.get("started_at"),
        completed_at=job_data.get("completed_at"),
        openai_cost_info=job_data.get("openai_cost_info")
    )

@app.post("/classify/batch", response_model=BatchClassificationResponse, deprecated=True, include_in_schema=False)
def classify_batch_products(request: BatchProductRequest):
    """
    [DEPRECATED] Clasificar múltiples productos en lote (síncrono)
    
    ⚠️ Este endpoint está deprecado. Use /classify/products en su lugar.
    
    Mantenido por compatibilidad hacia atrás.
    """
    batch_id = str(uuid.uuid4())
    results = []
    successful = 0
    failed = 0
    
    for idx, product in enumerate(request.products):
        try:
            result = classify(product.text, product.product_id)
            
            if 'error' not in result:
                results.append({
                    "index": idx,
                    "product_id": product.product_id,
                    "search_text": product.text,
                    "classification": result,
                    "status": "success",
                    "timestamp": datetime.now().isoformat()
                })
                successful += 1
            else:
                results.append({
                    "index": idx,
                    "product_id": product.product_id,
                    "search_text": product.text,
                    "error": result['error'],
                    "status": "failed",
                    "timestamp": datetime.now().isoformat()
                })
                failed += 1
                
        except Exception as e:
            results.append({
                "index": idx,
                "product_id": product.product_id,
                "search_text": product.text,
                "error": str(e),
                "status": "failed",
                "timestamp": datetime.now().isoformat()
            })
            failed += 1
    
    return BatchClassificationResponse(
        total=len(request.products),
        successful=successful,
        failed=failed,
        results=results,
        batch_id=batch_id
    )

def process_async_classification(products: List[ProductRequest], job_id: str, callback_url: Optional[str] = None):
    """
    Procesar clasificación asíncrona moderna con mejor tracking y manejo de errores
    """
    import time
    start_time = time.time()
    
    try:
        # Marcar job como iniciado
        background_jobs[job_id]["status"] = JobStatus.PROCESSING
        background_jobs[job_id]["started_at"] = datetime.now().isoformat()
        
        results = []
        successful = 0
        failed = 0
        
        # Variables para consolidar información de costos OpenAI
        total_prompt_tokens = 0
        total_completion_tokens = 0
        total_cost_usd = 0.0
        total_api_calls = 0
        model_used = None
        cost_info_sample = None
        
        for idx, product in enumerate(products):
            try:
                # Actualizar progreso en tiempo real
                current_progress = idx + 1
                percentage = (current_progress / len(products)) * 100
                background_jobs[job_id]["progress"] = {
                    "current": current_progress,
                    "total": len(products),
                    "percentage": round(percentage, 2)
                }
                
                # Procesar clasificación
                result = classify(product.text, product.product_id)
                
                if 'error' not in result:
                    # Clasificación exitosa
                    results.append({
                        "index": idx,
                        "product_id": product.product_id,
                        "search_text": product.text,
                        "classification": result,
                        "status": "success",
                        "timestamp": datetime.now().isoformat()
                    })
                    successful += 1
                    
                    # Consolidar información de costos si está disponible
                    if 'openai_cost_info' in result:
                        cost_info = result['openai_cost_info']
                        total_prompt_tokens += cost_info.get('usage', {}).get('prompt_tokens', 0)
                        total_completion_tokens += cost_info.get('usage', {}).get('completion_tokens', 0)
                        total_cost_usd += cost_info.get('cost_usd', {}).get('total', 0.0)
                        total_api_calls += cost_info.get('api_calls', 0)
                        if not model_used:
                            model_used = cost_info.get('model')
                            cost_info_sample = cost_info
                        
                else:
                    # Error en clasificación
                    results.append({
                        "index": idx,
                        "product_id": product.product_id,
                        "search_text": product.text,
                        "error": result['error'],
                        "status": "failed",
                        "timestamp": datetime.now().isoformat()
                    })
                    failed += 1
                    
            except Exception as e:
                # Error inesperado durante procesamiento
                results.append({
                    "index": idx,
                    "product_id": product.product_id,
                    "search_text": product.text,
                    "error": f"Error inesperado: {str(e)}",
                    "status": "failed",
                    "timestamp": datetime.now().isoformat()
                })
                failed += 1
        
        # Calcular tiempo total de procesamiento
        processing_time = time.time() - start_time
        
        # Crear información consolidada de costos OpenAI
        openai_cost_info = None
        if cost_info_sample and total_api_calls > 0:
            openai_cost_info = {
                "model": model_used,
                "usage": {
                    "prompt_tokens": total_prompt_tokens,
                    "completion_tokens": total_completion_tokens,
                    "total_tokens": total_prompt_tokens + total_completion_tokens
                },
                "cost_usd": {
                    "prompt": cost_info_sample['cost_usd']['prompt'] * (total_prompt_tokens / cost_info_sample['usage']['prompt_tokens']) if cost_info_sample['usage']['prompt_tokens'] > 0 else 0,
                    "completion": cost_info_sample['cost_usd']['completion'] * (total_completion_tokens / cost_info_sample['usage']['completion_tokens']) if cost_info_sample['usage']['completion_tokens'] > 0 else 0,
                    "total": total_cost_usd
                },
                "cost_breakdown": cost_info_sample.get('cost_breakdown', {}),
                "api_calls": total_api_calls
            }
        
        # Marcar job como completado exitosamente
        background_jobs[job_id].update({
            "status": JobStatus.COMPLETED,
            "completed_at": datetime.now().isoformat(),
            "results": results,
            "processing_time_seconds": round(processing_time, 3),
            "openai_cost_info": openai_cost_info
        })
        
        # TODO: Implementar callback notification si callback_url está presente
        if callback_url:
            # En una implementación completa aquí se haría HTTP POST al callback_url
            # con el job_id y estado final
            pass
            
    except Exception as e:
        # Error crítico durante todo el procesamiento
        background_jobs[job_id].update({
            "status": JobStatus.FAILED,
            "completed_at": datetime.now().isoformat(),
            "error_message": f"Error crítico en procesamiento: {str(e)}",
            "results": results if 'results' in locals() else []
        })

def process_batch_async(products: List[ProductRequest], job_id: str):
    """Procesar lote de productos en background"""
    background_jobs[job_id]["status"] = "processing"
    background_jobs[job_id]["started_at"] = datetime.now().isoformat()
    
    results = []
    successful = 0
    failed = 0
    
    for idx, product in enumerate(products):
        try:
            # Actualizar progreso
            background_jobs[job_id]["progress"] = {
                "current": idx + 1,
                "total": len(products),
                "percentage": ((idx + 1) / len(products)) * 100
            }
            
            result = classify(product.text, product.product_id)
            
            if 'error' not in result:
                results.append({
                    "index": idx,
                    "product_id": product.product_id,
                    "search_text": product.text,
                    "classification": result,
                    "status": "success",
                    "timestamp": datetime.now().isoformat()
                })
                successful += 1
            else:
                results.append({
                    "index": idx,
                    "product_id": product.product_id,
                    "search_text": product.text,
                    "error": result['error'],
                    "status": "failed", 
                    "timestamp": datetime.now().isoformat()
                })
                failed += 1
                
        except Exception as e:
            results.append({
                "index": idx,
                "product_id": product.product_id,
                "search_text": product.text,
                "error": str(e),
                "status": "failed",
                "timestamp": datetime.now().isoformat()
            })
            failed += 1
    
    # Completar job
    background_jobs[job_id].update({
        "status": "completed",
        "completed_at": datetime.now().isoformat(),
        "results": {
            "total": len(products),
            "successful": successful,
            "failed": failed,
            "data": results
        }
    })

@app.post("/classify/batch/async", deprecated=True, include_in_schema=False)
def classify_batch_async(request: BatchProductRequest, background_tasks: BackgroundTasks):
    """
    [DEPRECATED] Clasificar múltiples productos en lote (asíncrono)
    
    ⚠️ Este endpoint está deprecado. Use /classify/products para clasificación síncrona.
    
    Mantenido por compatibilidad hacia atrás.
    """
    job_id = str(uuid.uuid4())
    
    # Inicializar job
    background_jobs[job_id] = {
        "status": "queued",
        "created_at": datetime.now().isoformat(),
        "total_products": len(request.products),
        "progress": {"current": 0, "total": len(request.products), "percentage": 0}
    }
    
    # Agregar tarea en background
    background_tasks.add_task(process_batch_async, request.products, job_id)
    
    return {
        "job_id": job_id,
        "status": "queued",
        "message": "Procesamiento iniciado en background",
        "total_products": len(request.products),
        "check_status_url": f"/jobs/{job_id}"
    }

@app.get("/jobs/{job_id}")
def get_job_status(job_id: str):
    """Obtener estado de un trabajo en background"""
    if job_id not in background_jobs:
        raise HTTPException(status_code=404, detail="Job no encontrado")
    
    return background_jobs[job_id]

# Endpoints adicionales de utilidad
@app.get("/stats")
def get_api_stats():
    """Estadísticas de uso de la API"""
    total_jobs = len(background_jobs)
    completed = sum(1 for job in background_jobs.values() if job.get("status") == "completed")
    processing = sum(1 for job in background_jobs.values() if job.get("status") == "processing")
    
    return {
        "total_background_jobs": total_jobs,
        "completed_jobs": completed,
        "processing_jobs": processing,
        "api_uptime": "N/A",  # En producción calcular uptime real
        "timestamp": datetime.now().isoformat()
    }

# Funciones de exportación
def create_csv_export(results_data: List[Dict], filename: str = None) -> str:
    """
    Crea un archivo CSV con los resultados de clasificación
    
    Args:
        results_data: Lista con resultados de clasificación
        filename: Nombre del archivo (opcional)
        
    Returns:
        str: Ruta completa del archivo creado
    """
    ensure_export_structure()
    
    if not filename:
        output_path = get_full_export_path("api_export", "csv")
    else:
        output_path = get_full_export_path(filename.replace('.csv', ''), "csv")
    
    with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
        fieldnames = [
            'product_id',
            'product_description', 
            'skos_category',
            'skos_notation',
            'skos_uri',
            'confidence_score',
            'classification_timestamp',
            'status'
        ]
        
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        
        for item in results_data:
            if item.get('status') == 'success' and 'classification' in item:
                classification = item['classification']
                
                # Manejar prefLabel que puede ser un diccionario
                pref_label = classification.get('prefLabel', '')
                if isinstance(pref_label, dict):
                    pref_label = pref_label.get('es', pref_label.get('en', str(pref_label)))
                
                row = {
                    'product_id': item.get('product_id', ''),
                    'product_description': item.get('search_text', ''),
                    'skos_category': pref_label,
                    'skos_notation': classification.get('notation', ''),
                    'skos_uri': classification.get('concept_uri', ''),
                    'confidence_score': classification.get('confidence', 0),
                    'classification_timestamp': item.get('timestamp', ''),
                    'status': 'success'
                }
            else:
                row = {
                    'product_id': item.get('product_id', ''),
                    'product_description': item.get('search_text', ''),
                    'skos_category': f"ERROR: {item.get('error', 'Unknown error')}",
                    'skos_notation': '',
                    'skos_uri': '',
                    'confidence_score': 0,
                    'classification_timestamp': item.get('timestamp', ''),
                    'status': 'error'
                }
            writer.writerow(row)
    
    return str(output_path)

def create_excel_export(results_data: List[Dict], filename: str = None) -> str:
    """
    Crea un archivo Excel con los resultados de clasificación
    
    Args:
        results_data: Lista con resultados de clasificación
        filename: Nombre del archivo (opcional)
        
    Returns:
        str: Ruta completa del archivo creado
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500, 
            detail="openpyxl no está instalado. Ejecute: pip install openpyxl"
        )
    
    ensure_export_structure()
    
    if not filename:
        output_path = get_full_export_path("api_export", "excel")
    else:
        output_path = get_full_export_path(filename.replace('.xlsx', ''), "excel")
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos Clasificados"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "ID Producto",
        "Descripción del Producto", 
        "Categoría SKOS",
        "Notación SKOS",
        "URI Concepto",
        "Nivel de Confianza",
        "Timestamp Clasificación",
        "Estado"
    ]
    
    # Escribir headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Escribir datos
    for row_idx, item in enumerate(results_data, 2):
        if item.get('status') == 'success' and 'classification' in item:
            classification = item['classification']
            
            # Manejar prefLabel que puede ser un diccionario
            pref_label = classification.get('prefLabel', '')
            if isinstance(pref_label, dict):
                pref_label = pref_label.get('es', pref_label.get('en', str(pref_label)))
            
            data = [
                item.get('product_id', ''),
                item.get('search_text', ''),
                pref_label,
                classification.get('notation', ''),
                classification.get('concept_uri', ''),
                classification.get('confidence', 0),
                item.get('timestamp', ''),
                'success'
            ]
        else:
            data = [
                item.get('product_id', ''),
                item.get('search_text', ''),
                f"ERROR: {item.get('error', 'Unknown error')}",
                '',
                '',
                0,
                item.get('timestamp', ''),
                'error'
            ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = border
    
    # Ajustar ancho de columnas
    column_widths = [15, 40, 25, 15, 50, 15, 20, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Guardar archivo
    wb.save(output_path)
    
    return str(output_path)

# Endpoints de exportación
@app.post("/export/csv", response_model=ExportResponse)
def export_to_csv_endpoint(request: ExportRequest):
    """Exportar productos clasificados a CSV y generar URL de descarga"""
    try:
        # Clasificar productos
        results = []
        successful = 0
        failed = 0
        
        for idx, product in enumerate(request.products):
            try:
                result = classify(product.text, product.product_id)
                
                if 'error' not in result:
                    results.append({
                        "index": idx,
                        "product_id": product.product_id,
                        "search_text": product.text,
                        "classification": result,
                        "status": "success",
                        "timestamp": datetime.now().isoformat()
                    })
                    successful += 1
                else:
                    results.append({
                        "index": idx,
                        "product_id": product.product_id,
                        "search_text": product.text,
                        "error": result['error'],
                        "status": "failed",
                        "timestamp": datetime.now().isoformat()
                    })
                    failed += 1
                    
            except Exception as e:
                results.append({
                    "index": idx,
                    "product_id": product.product_id,
                    "search_text": product.text,
                    "error": str(e),
                    "status": "failed",
                    "timestamp": datetime.now().isoformat()
                })
                failed += 1
        
        # Crear archivo CSV
        file_path = create_csv_export(results, request.filename)
        filename = Path(file_path).name
        
        # Obtener tamaño del archivo
        file_size = Path(file_path).stat().st_size
        
        return ExportResponse(
            status="success",
            filename=filename,
            download_url=f"/download/{filename}",
            total_products=len(request.products),
            successful=successful,
            failed=failed,
            file_size_bytes=file_size,
            timestamp=datetime.now().isoformat()
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/export/excel", response_model=ExportResponse)
def export_to_excel_endpoint(request: ExportRequest):
    """Exportar productos clasificados a Excel y generar URL de descarga"""
    try:
        # Clasificar productos
        results = []
        successful = 0
        failed = 0
        
        for idx, product in enumerate(request.products):
            try:
                result = classify(product.text, product.product_id)
                
                if 'error' not in result:
                    results.append({
                        "index": idx,
                        "product_id": product.product_id,
                        "search_text": product.text,
                        "classification": result,
                        "status": "success",
                        "timestamp": datetime.now().isoformat()
                    })
                    successful += 1
                else:
                    results.append({
                        "index": idx,
                        "product_id": product.product_id,
                        "search_text": product.text,
                        "error": result['error'],
                        "status": "failed",
                        "timestamp": datetime.now().isoformat()
                    })
                    failed += 1
                    
            except Exception as e:
                results.append({
                    "index": idx,
                    "product_id": product.product_id,
                    "search_text": product.text,
                    "error": str(e),
                    "status": "failed",
                    "timestamp": datetime.now().isoformat()
                })
                failed += 1
        
        # Crear archivo Excel
        file_path = create_excel_export(results, request.filename)
        filename = Path(file_path).name
        
        # Obtener tamaño del archivo
        file_size = Path(file_path).stat().st_size
        
        return ExportResponse(
            status="success",
            filename=filename,
            download_url=f"/download/{filename}",
            total_products=len(request.products),
            successful=successful,
            failed=failed,
            file_size_bytes=file_size,
            timestamp=datetime.now().isoformat()
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/download/{filename}")
def download_file(filename: str):
    """Descargar archivo exportado"""
    # Buscar archivo en las subcarpetas de exports
    file_found = None
    
    # Buscar en todas las subcarpetas de exports
    for root, dirs, files in os.walk(EXPORTS_BASE_DIR):
        if filename in files:
            file_found = Path(root) / filename
            break
    
    if not file_found or not file_found.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    
    # Determinar content type
    content_type = "application/octet-stream"
    if filename.endswith('.csv'):
        content_type = "text/csv"
    elif filename.endswith('.xlsx'):
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif filename.endswith('.json'):
        content_type = "application/json"
    
    return FileResponse(
        path=str(file_found),
        filename=filename,
        media_type=content_type
    )

# === ENDPOINTS MEJORADOS PARA MANEJO DE NO CLASIFICABLES ===

@app.post("/classify/enhanced", response_model=Dict[str, Any])
def classify_single_product_enhanced(
    request: ProductRequest,
    taxonomy: Optional[str] = Query(None, description="ID de taxonomía específica a usar")
):
    """
    🎯 Clasificación mejorada con manejo inteligente de productos no clasificables
    
    **Mejoras v3.1:**
    - Detección automática de incompatibilidad dominio/taxonomía
    - Respuestas estructuradas para productos no clasificables
    - Sugerencias específicas para resolver problemas
    - Análisis de calidad del input
    
    En lugar de devolver errores genéricos, proporciona información útil sobre:
    - Por qué no se pudo clasificar el producto
    - Qué dominio detectó para el producto
    - Qué taxonomías serían más apropiadas
    - Cómo mejorar la descripción del producto
    """
    try:
        # Usar taxonomía por defecto si no se especifica
        from utils.taxonomy_config import validate_taxonomy_id, get_taxonomy_info
        
        validated_taxonomy = validate_taxonomy_id(taxonomy)
        taxonomy_info = get_taxonomy_info(validated_taxonomy)
        
        # Ejecutar clasificación base
        result = classify(request.text, request.product_id, validated_taxonomy)
        
        # Si hay error, aplicar manejo mejorado
        if 'error' in result:
            enhanced_result = enhance_classification_error_handling(
                original_result=result,
                text=request.text,
                product_id=request.product_id,
                taxonomy_id=validated_taxonomy
            )
            
            # Agregar información de taxonomía
            enhanced_result['taxonomy_used'] = {
                'id': validated_taxonomy,
                'name': taxonomy_info.get('name', 'Unknown'),
                'is_default': taxonomy is None
            }
            
            return enhanced_result
        
        # Si clasificación exitosa, agregar info de taxonomía y devolver
        result['taxonomy_used'] = {
            'id': validated_taxonomy,
            'name': taxonomy_info.get('name', 'Unknown'),
            'is_default': taxonomy is None
        }
        
        return {
            "classification_result": "success",
            "classification": result,
            "taxonomy_used": result['taxonomy_used']
        }
        
    except Exception as e:
        return {
            "classification_result": "error",
            "error": {
                "type": "system_error",
                "message": str(e),
                "timestamp": datetime.now().isoformat()
            },
            "product_id": request.product_id,
            "original_text": request.text
        }

@app.post("/classify/products/enhanced", response_model=Dict[str, Any])
def classify_products_enhanced(
    request: UnifiedProductRequest,
    taxonomy: Optional[str] = Query(None, description="ID de taxonomía específica a usar")
):
    """
    🚀 Clasificación en lotes mejorada con análisis detallado de productos no clasificables
    
    Proporciona estadísticas detalladas sobre:
    - Productos clasificados exitosamente
    - Productos no clasificables con análisis de causa
    - Incompatibilidades de dominio detectadas
    - Sugerencias de mejora agregadas
    """
    import time
    start_time = time.time()
    
    try:
        from utils.taxonomy_config import validate_taxonomy_id, get_taxonomy_info
        
        validated_taxonomy = validate_taxonomy_id(taxonomy)
        taxonomy_info = get_taxonomy_info(validated_taxonomy)
        
        results = []
        successful = 0
        failed = 0
        not_classifiable = 0
        domain_mismatches = 0
        
        for idx, product in enumerate(request.products):
            # Ejecutar clasificación
            result = classify(product.text, product.product_id, validated_taxonomy)
            
            if 'error' not in result:
                # Clasificación exitosa
                results.append({
                    "index": idx,
                    "product_id": product.product_id,
                    "search_text": product.text,
                    "classification": result,
                    "status": "success",
                    "timestamp": datetime.now().isoformat()
                })
                successful += 1
            else:
                # Aplicar manejo mejorado de errores
                enhanced_result = enhance_classification_error_handling(
                    original_result=result,
                    text=product.text,
                    product_id=product.product_id,
                    taxonomy_id=validated_taxonomy
                )
                
                # Determinar tipo de fallo
                if enhanced_result.get('classification_result') == 'not_classifiable':
                    not_classifiable += 1
                    if enhanced_result.get('reason') == 'domain_mismatch':
                        domain_mismatches += 1
                else:
                    failed += 1
                
                results.append({
                    "index": idx,
                    "product_id": product.product_id,
                    "search_text": product.text,
                    "enhanced_analysis": enhanced_result,
                    "status": "not_classifiable" if enhanced_result.get('classification_result') == 'not_classifiable' else "error",
                    "timestamp": datetime.now().isoformat()
                })
        
        processing_time = time.time() - start_time
        
        # Estadísticas agregadas
        total_processed = len(request.products)
        
        return {
            "total": total_processed,
            "successful": successful,
            "not_classifiable": not_classifiable,
            "failed": failed,
            "domain_mismatches": domain_mismatches,
            "results": results,
            "processing_summary": {
                "success_rate": (successful / total_processed) * 100 if total_processed > 0 else 0,
                "not_classifiable_rate": (not_classifiable / total_processed) * 100 if total_processed > 0 else 0,
                "domain_mismatch_rate": (domain_mismatches / total_processed) * 100 if total_processed > 0 else 0,
                "processing_time_seconds": processing_time,
                "average_time_per_product": processing_time / total_processed if total_processed > 0 else 0
            },
            "taxonomy_used": {
                "id": validated_taxonomy,
                "name": taxonomy_info.get('name', 'Unknown'),
                "is_default": taxonomy is None
            },
            "recommendations": {
                "overall": f"Tasa de éxito del {(successful / total_processed) * 100:.1f}%" if total_processed > 0 else "Sin productos procesados",
                "suggested_actions": _generate_batch_recommendations(successful, not_classifiable, domain_mismatches, total_processed)
            },
            "timestamp": datetime.now().isoformat()
        }
        
    except Exception as e:
        return {
            "error": {
                "type": "batch_processing_error",
                "message": str(e),
                "timestamp": datetime.now().isoformat()
            },
            "total": len(request.products),
            "processed": 0
        }

def _generate_batch_recommendations(successful: int, not_classifiable: int, domain_mismatches: int, total: int) -> List[str]:
    """Generar recomendaciones para el lote procesado"""
    recommendations = []
    
    success_rate = (successful / total) * 100 if total > 0 else 0
    mismatch_rate = (domain_mismatches / total) * 100 if total > 0 else 0
    
    if success_rate > 80:
        recommendations.append("Excelente tasa de éxito, taxonomía bien alineada")
    elif success_rate > 50:
        recommendations.append("Tasa de éxito moderada, revisar productos no clasificables")
    else:
        recommendations.append("Baja tasa de éxito, considerar cambio de taxonomía")
    
    if mismatch_rate > 30:
        recommendations.append("Alto porcentaje de incompatibilidad de dominio, considerar taxonomía multi-dominio")
    
    if not_classifiable > successful:
        recommendations.append("Más productos no clasificables que exitosos, revisar compatibilidad taxonomía-catálogo")
    
    return recommendations

if __name__ == "__main__":
    import uvicorn
    print("🚀 Iniciando SKOS Classification API...")
    print("📖 Documentación: http://localhost:8000/docs")
    print("🔗 API: http://localhost:8000")
    uvicorn.run(app, host="0.0.0.0", port=8000)